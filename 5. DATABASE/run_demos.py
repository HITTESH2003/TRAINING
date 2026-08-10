"""
Module 5: Database — a guided tour of the Postgres/pgvector database Module 4
already built and populated, run for real against the live container.

Runs four demos against it:
  1. Distance operators (<->, <=>, <#>) -- do they actually rank differently?
  2. Metadata-filtered vector search -- does restricting to one document change the answer?
  3. Full-text vs vector search -- same queries, two different search engines, compared
  4. Index benchmark -- a separate synthetic table, seq scan vs HNSW vs IVFFlat, real EXPLAIN ANALYZE

Requires Module 4's Postgres container running with data already loaded:
  cd "../4. EMBEDDING AND DATABASE" && docker compose up -d
  "../1. OCR/.venv/bin/python" load_and_store.py

Usage:
  python run_demos.py
  python run_demos.py --skip-index-benchmark   # skip the slow 200k-row build
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from db_features import connection, fulltext_search, index_benchmark, metadata_search, operators_demo

REPORT_PATH = Path(__file__).resolve().parent / "database_features_report.xlsx"

OPERATOR_QUERIES = [
    "What percentage did illegal border crossings drop by?",
    "Which countries are NASA's international partners on the space station?",
    "How is efficiency calculated in the regional operations report?",
]

METADATA_FILTER_DEMO = {
    "query": "important developments and changes",
    "doc_name": "nasa_iss_factsheet",
}

FULLTEXT_VS_VECTOR_QUERIES = [
    "92 percent",
    "Houston Texas",
    "orbital laboratory",
    "how much did border crossings decrease",
    "an orbiting laboratory where things float",
]


def run_operator_demo(conn) -> list[dict]:
    print("\n=== 1. Distance operators: <->  vs  <=>  vs  <#> ===")
    results = []
    for query in OPERATOR_QUERIES:
        result = operators_demo.run(conn, query)
        operators_demo.print_result(result)
        results.append(result)
    return results


def run_metadata_demo(conn) -> dict:
    print("\n=== 2. Metadata-filtered vector search ===")
    result = metadata_search.compare_filtered_vs_unfiltered(conn, METADATA_FILTER_DEMO["query"], METADATA_FILTER_DEMO["doc_name"])
    print(f"Query: {result['query']!r}  filtered to doc_name={result['doc_name_filter']!r}")
    print(f"  unfiltered top-1: {result['unfiltered_top1']['chunk_id']}")
    print(f"  filtered   top-1: {result['filtered_top1']['chunk_id']}")
    print(f"  -> filter changed the top result: {result['top1_changed_by_filter']}")
    return result


def run_fulltext_demo(conn) -> list[dict]:
    print("\n=== 3. Full-text search vs vector search ===")
    fulltext_search.ensure_fulltext_index(conn)
    results = []
    for query in FULLTEXT_VS_VECTOR_QUERIES:
        result = fulltext_search.compare(conn, query)
        ft = [(r["chunk_id"], round(r["rank"], 3)) for r in result["fulltext_results"]]
        vec = [(r["chunk_id"], round(r["distance"], 3)) for r in result["vector_results"]]
        print(f"Query: {query!r}")
        print(f"  fulltext: {ft or '(no matches)'}")
        print(f"  vector:   {vec}")
        results.append(result)
    return results


def run_index_benchmark(conn) -> dict:
    print("\n=== 4. Index benchmark (synthetic data, not real chunks) ===")
    index_benchmark.build_benchmark_table(conn)
    results = index_benchmark.benchmark(conn)
    for method, stats in results.items():
        print(f"  {method:10s} scan_type={stats['scan_type']:12s} wall_clock_ms={stats['wall_clock_ms']:8.2f}", end="")
        if "build_time_s" in stats:
            print(f"  build_time_s={stats['build_time_s']}", end="")
        print()
    return results


def _write_report(operator_results, metadata_result, fulltext_results, index_results) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Distance Operators"
    ws.append(["query", "l2_top1", "cosine_top1", "inner_product_top1", "same_ranking"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in operator_results:
        ws.append([r["query"], r["rankings"]["l2"][0], r["rankings"]["cosine"][0], r["rankings"]["inner_product"][0], r["same_ranking_across_operators"]])
    for i, w in enumerate([55, 45, 45, 45, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Metadata Filtering")
    ws2.append(["query", "doc_name_filter", "unfiltered_top1", "filtered_top1", "top1_changed"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.append(
        [
            metadata_result["query"],
            metadata_result["doc_name_filter"],
            metadata_result["unfiltered_top1"]["chunk_id"],
            metadata_result["filtered_top1"]["chunk_id"],
            metadata_result["top1_changed_by_filter"],
        ]
    )
    for i, w in enumerate([40, 25, 45, 45, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Fulltext vs Vector")
    ws3.append(["query", "fulltext_top1", "vector_top1", "fulltext_found_anything"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for r in fulltext_results:
        ft_top1 = r["fulltext_results"][0]["chunk_id"] if r["fulltext_results"] else "(none)"
        vec_top1 = r["vector_results"][0]["chunk_id"] if r["vector_results"] else "(none)"
        ws3.append([r["query"], ft_top1, vec_top1, bool(r["fulltext_results"])])
    for i, w in enumerate([45, 45, 45, 20], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws4 = wb.create_sheet("Index Benchmark")
    ws4.append(["method", "scan_type", "wall_clock_ms", "planner_time_ms", "build_time_s"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for method, stats in index_results.items():
        ws4.append([method, stats["scan_type"], stats["wall_clock_ms"], stats["planner_time_ms"], stats.get("build_time_s", "")])
    for i, w in enumerate([12, 14, 16, 16, 14], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(REPORT_PATH)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Module 5's database feature demos against Module 4's live Postgres container")
    parser.add_argument("--skip-index-benchmark", action="store_true", help="skip the slow synthetic 200k-row index benchmark")
    args = parser.parse_args()

    conn = connection.connect()
    n = connection.require_chunks_table(conn)
    print(f"Connected. `chunks` table has {n} rows (from Module 4).")

    operator_results = run_operator_demo(conn)
    metadata_result = run_metadata_demo(conn)
    fulltext_results = run_fulltext_demo(conn)
    index_results = {} if args.skip_index_benchmark else run_index_benchmark(conn)

    conn.close()

    if index_results:
        _write_report(operator_results, metadata_result, fulltext_results, index_results)
        print(f"\nWrote {REPORT_PATH}")
    else:
        print("\nSkipped index benchmark -- report not written (needs all four sections).")


if __name__ == "__main__":
    main()
