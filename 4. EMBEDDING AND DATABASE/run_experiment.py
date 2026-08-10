"""
Module 4 experiment: how much does shrinking the embedding dimension actually
cost you at retrieval time?

For every query in embedding/queries.py, at every dimension in
embedding/config.py's TEST_DIMENSIONS, this runs a REAL pgvector nearest-
neighbor search (SQL, not numpy) against the chunks table Module 4 already
populated, and checks where the known-correct chunk lands in the ranking.

Usage:
  python run_experiment.py
"""

from __future__ import annotations

import statistics
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from embedding import config, db, nomic_embed
from embedding.queries import QUERIES

TOP_K = 5
REPORT_PATH = Path(__file__).resolve().parent / "dimension_accuracy_report.xlsx"


def run() -> dict[int, dict]:
    conn = db.connect()
    total_chunks = db.chunk_count(conn)
    if total_chunks == 0:
        raise SystemExit("The chunks table is empty -- run load_and_store.py first.")
    print(f"Running {len(QUERIES)} queries against {total_chunks} stored chunks, at dimensions {config.TEST_DIMENSIONS} ...")

    detail_rows = []
    dim_results: dict[int, dict] = {}

    for dim in config.TEST_DIMENSIONS:
        reciprocal_ranks = []
        hits_at_1 = 0
        hits_at_3 = 0
        distances_to_correct = []

        for item in QUERIES:
            query_embeddings = nomic_embed.embed_query_all_dims(item["query"])
            query_vec = query_embeddings[dim]
            results = db.nearest_neighbors(conn, dim, query_vec, TOP_K)

            rank = None
            for i, row in enumerate(results, start=1):
                if row["chunk_id"] == item["expected_chunk_id"]:
                    rank = i
                    break

            if rank == 1:
                hits_at_1 += 1
            if rank is not None and rank <= 3:
                hits_at_3 += 1
            reciprocal_ranks.append(1.0 / rank if rank else 0.0)

            expected_row = next((r for r in results if r["chunk_id"] == item["expected_chunk_id"]), None)
            if expected_row is None:
                # expected chunk wasn't even in the top TOP_K -- look up its distance directly
                own = db.nearest_neighbors(conn, dim, query_vec, total_chunks)
                expected_row = next(r for r in own if r["chunk_id"] == item["expected_chunk_id"])
            distances_to_correct.append(expected_row["distance"])

            detail_rows.append(
                {
                    "dimension": dim,
                    "query": item["query"],
                    "expected_chunk_id": item["expected_chunk_id"],
                    "rank_of_expected": rank if rank else f">{TOP_K}",
                    "top1_chunk_id": results[0]["chunk_id"],
                    "top1_distance": round(results[0]["distance"], 4),
                    "distance_to_expected": round(expected_row["distance"], 4),
                }
            )

        n = len(QUERIES)
        dim_results[dim] = {
            "dimension": dim,
            "recall_at_1": round(hits_at_1 / n, 3),
            "recall_at_3": round(hits_at_3 / n, 3),
            "mean_reciprocal_rank": round(statistics.mean(reciprocal_ranks), 3),
            "avg_distance_to_correct": round(statistics.mean(distances_to_correct), 4),
        }
        print(
            f"  dim={dim:4d}  recall@1={dim_results[dim]['recall_at_1']:.2f}  "
            f"recall@3={dim_results[dim]['recall_at_3']:.2f}  "
            f"MRR={dim_results[dim]['mean_reciprocal_rank']:.2f}  "
            f"avg_dist_to_correct={dim_results[dim]['avg_distance_to_correct']:.4f}"
        )

    conn.close()
    _write_report(dim_results, detail_rows)
    print(f"\nWrote {REPORT_PATH}")
    return dim_results


def _write_report(dim_results: dict[int, dict], detail_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Accuracy by Dimension"
    cols = ["dimension", "recall_at_1", "recall_at_3", "mean_reciprocal_rank", "avg_distance_to_correct"]
    summary_ws.append(cols)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    for dim in config.TEST_DIMENSIONS:
        summary_ws.append([dim_results[dim][c] for c in cols])
    for i, w in enumerate([12, 13, 13, 20, 22], start=1):
        summary_ws.column_dimensions[get_column_letter(i)].width = w

    detail_ws = wb.create_sheet("Per-Query Detail")
    detail_cols = ["dimension", "query", "expected_chunk_id", "rank_of_expected", "top1_chunk_id", "top1_distance", "distance_to_expected"]
    detail_ws.append(detail_cols)
    for cell in detail_ws[1]:
        cell.font = Font(bold=True)
    for row in detail_rows:
        detail_ws.append([row[c] for c in detail_cols])
    for i, w in enumerate([12, 55, 45, 16, 45, 14, 18], start=1):
        detail_ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(REPORT_PATH)


if __name__ == "__main__":
    run()
