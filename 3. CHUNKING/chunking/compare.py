"""Runs every chunking strategy over one document and produces a side-by-side
comparison: chunk count, token-size spread, and concrete evidence of the
failure modes each strategy is prone to (split tables, tiny orphan chunks)."""

from __future__ import annotations

import json
import statistics
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import DEFAULT_MAX_TOKENS, DEFAULT_MIN_TOKENS, DEFAULT_SIMILARITY_THRESHOLD
from .hierarchical import chunk_hierarchical
from .layout_blocks import load_layout_blocks, regions_path_for
from .recursive import chunk_recursive
from .semantic import chunk_semantic

STRATEGIES = ("recursive", "semantic", "hierarchical")


def run_all_strategies(
    markdown_path: Path,
    max_tokens: int = DEFAULT_MAX_TOKENS,
    min_tokens: int = DEFAULT_MIN_TOKENS,
    similarity_threshold: float = DEFAULT_SIMILARITY_THRESHOLD,
) -> dict[str, list[dict]]:
    doc_name = markdown_path.stem
    raw_text = markdown_path.read_text(encoding="utf-8")

    # semantic and hierarchical need real layout classes (title/table/figure/...),
    # not a re-guess from markdown syntax -- see layout_blocks.py. recursive
    # deliberately works on raw_text with no structure at all.
    regions_json_path = regions_path_for(markdown_path.parent)
    if not regions_json_path.exists():
        raise FileNotFoundError(
            f"{regions_json_path} not found -- re-run Module 1 (1. OCR) on this document; "
            "regions.json is required for semantic/hierarchical chunking."
        )
    blocks = load_layout_blocks(regions_json_path)

    return {
        "recursive": chunk_recursive(raw_text, doc_name, max_tokens=max_tokens),
        "semantic": chunk_semantic(blocks, doc_name, max_tokens=max_tokens, similarity_threshold=similarity_threshold),
        "hierarchical": chunk_hierarchical(blocks, doc_name, max_tokens=max_tokens, min_tokens=min_tokens),
    }


def _strategy_stats(doc_name: str, strategy: str, chunks: list[dict], min_tokens: int) -> dict:
    token_counts = [c["token_count"] for c in chunks] or [0]
    broken = sum(1 for c in chunks if "table_split" in c.get("notes", []))
    oversized = sum(1 for c in chunks if "oversized_atomic" in c.get("notes", []) or "oversized_block" in c.get("notes", []))
    tiny = sum(1 for c in chunks if c["token_count"] < min_tokens and "oversized_atomic" not in c.get("notes", []))
    boundary_similarities = [c["boundary_similarity"] for c in chunks if c.get("boundary_similarity") is not None]
    return {
        "document": doc_name,
        "strategy": strategy,
        "chunk_count": len(chunks),
        "avg_tokens": round(statistics.mean(token_counts), 1),
        "min_tokens": min(token_counts),
        "max_tokens": max(token_counts),
        "stdev_tokens": round(statistics.pstdev(token_counts), 1) if len(token_counts) > 1 else 0.0,
        "tiny_chunks_below_min": tiny,
        "table_split_count": broken,
        "oversized_atomic_count": oversized,
        "avg_boundary_similarity": round(statistics.mean(boundary_similarities), 3) if boundary_similarities else None,
    }


def write_json_outputs(doc_name: str, results: dict[str, list[dict]], output_dir: Path) -> None:
    doc_dir = output_dir / doc_name
    doc_dir.mkdir(parents=True, exist_ok=True)
    for strategy, chunks in results.items():
        (doc_dir / f"{strategy}.json").write_text(json.dumps(chunks, indent=2), encoding="utf-8")


def build_comparison_report(all_results: dict[str, dict[str, list[dict]]], report_path: Path, min_tokens: int) -> None:
    """all_results: {doc_name: {strategy: [chunk, ...]}}"""
    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Comparison"
    summary_cols = [
        "document", "strategy", "chunk_count", "avg_tokens", "min_tokens", "max_tokens",
        "stdev_tokens", "tiny_chunks_below_min", "table_split_count", "oversized_atomic_count",
        "avg_boundary_similarity",
    ]
    summary_ws.append(summary_cols)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    detail_ws = wb.create_sheet("Chunks Detail")
    detail_cols = ["document", "strategy", "chunk_id", "token_count", "section_path", "pages", "block_types", "boundary_similarity", "notes", "text_preview"]
    detail_ws.append(detail_cols)
    for cell in detail_ws[1]:
        cell.font = Font(bold=True)

    for doc_name, strategies in all_results.items():
        for strategy, chunks in strategies.items():
            stats = _strategy_stats(doc_name, strategy, chunks, min_tokens)
            summary_ws.append([stats[c] for c in summary_cols])
            for chunk in chunks:
                preview = chunk["text"][:150].replace("\n", " ")
                detail_ws.append(
                    [
                        doc_name,
                        strategy,
                        chunk["chunk_id"],
                        chunk["token_count"],
                        ", ".join(chunk["section_path"]) if chunk.get("section_path") else "",
                        ", ".join(str(p) for p in chunk["pages"]) if chunk.get("pages") else "",
                        ", ".join(chunk["block_types"]) if chunk.get("block_types") else "",
                        chunk.get("boundary_similarity") if chunk.get("boundary_similarity") is not None else "",
                        ", ".join(chunk.get("notes", [])),
                        preview,
                    ]
                )

    for i, width in enumerate([28, 14, 12, 11, 11, 11, 12, 18, 16, 18, 20], start=1):
        summary_ws.column_dimensions[get_column_letter(i)].width = width
    for i, width in enumerate([28, 14, 32, 12, 30, 12, 20, 16, 22, 70], start=1):
        detail_ws.column_dimensions[get_column_letter(i)].width = width

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)
