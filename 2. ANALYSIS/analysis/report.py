"""Builds the Excel workbook: one summary row per document, plus detail
sheets for artifacts and low-confidence regions worth a manual second look."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SUMMARY_COLUMNS = [
    ("document", 28),
    ("page_count", 11),
    ("word_count", 11),
    ("token_count", 12),
    ("tokens_per_page", 14),
    ("images_in_body", 13),
    ("tables_in_body", 13),
    ("logo_count", 11),
    ("signature_count", 13),
    ("stamp_seal_count", 14),
    ("other_artifact_count", 16),
    ("low_confidence_count", 17),
    ("dropped_region_count", 17),
    ("malformed_table_count", 18),
    ("title", 30),
    ("author", 20),
    ("creation_date", 20),
]


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_report(results: list[dict], report_path: Path) -> None:
    if not results:
        raise ValueError("No documents to report on")

    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    col_names = [c for c, _ in SUMMARY_COLUMNS]
    _write_header(summary_ws, col_names)
    for r in results:
        summary_ws.append([r[c] for c in col_names])
    _autosize(summary_ws, [w for _, w in SUMMARY_COLUMNS])

    artifacts_ws = wb.create_sheet("Artifacts Detail")
    _write_header(artifacts_ws, ["document", "page", "type", "description", "bbox"])
    for r in results:
        for a in r["artifacts"]:
            artifacts_ws.append([r["document"], a["page"], a["type"], a["description"], str(a["bbox"])])
    _autosize(artifacts_ws, [28, 8, 14, 70, 30])

    low_conf_ws = wb.create_sheet("Low Confidence Regions")
    _write_header(low_conf_ws, ["document", "page", "class", "score", "bbox"])
    for r in results:
        for lc in r["low_confidence_regions"]:
            low_conf_ws.append([r["document"], lc["page"], lc["class"], lc["score"], str(lc["bbox"])])
    _autosize(low_conf_ws, [28, 8, 16, 10, 30])

    dropped_ws = wb.create_sheet("Dropped Regions")
    _write_header(dropped_ws, ["document", "page", "class", "score", "bbox"])
    for r in results:
        for dr in r["dropped_regions"]:
            dropped_ws.append([r["document"], dr["page"], dr["class"], dr["score"], str(dr["bbox"])])
    _autosize(dropped_ws, [28, 8, 16, 10, 30])

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)
